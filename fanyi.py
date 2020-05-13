# -*- coding: utf-8 -*-
import openpyxl,requests,json
from urllib import request
from urllib import parse

wb = openpyxl.load_workbook('123.xlsx')
# res = ws.cell(row=1, column=3).value
# print(res)

def main(i,j):
        res = wb[wb.get_sheet_names()[0]].cell(row=i, column=j).value
            return res

        def translate(content):
             
                Request_URL = 'http://fanyi.youdao.com/translate?smartresult=dict&smartresult=rule'
                    Form_Data = {}
                        Form_Data['i'] = content
                            Form_Data['from'] = 'AUTO'
                                Form_Data['to'] = 'AUTO'
                                    Form_Data['smartresult'] = 'dict'
                                        Form_Data['client'] = 'fanyideskweb'
                                            Form_Data['doctype'] = 'json'
                                                Form_Data['version'] = '2.1'
                                                    Form_Data['keyfrom'] = 'fanyi.web'
                                                        Form_Data['action'] = 'FY_BY_REALTIME'
                                                            Form_Data['typoResult'] = 'false'
                                                                data = parse.urlencode(Form_Data).encode('utf-8')
                                                                    response = request.urlopen(Request_URL, data)
                                                                        html = response.read().decode('utf-8')
                                                                            translate_results = json.loads(html)
                                                                                translate_results = translate_results['translateResult'][0][0]['tgt']
                                                                                    return  translate_results

                                                                                if __name__ == '__main__':
                                                                                        for i in range(1,226):
                                                                                                    for j in range(1,15):
                                                                                                                    print(translate(main(i,j)))# -*- coding: utf-8 -*-
                                                                                                                    import openpyxl,requests,json
                                                                                                                    from urllib import request
                                                                                                                    from urllib import parse

                                                                                                                    wb = openpyxl.load_workbook('123.xlsx')
                                                                                                                    # res = ws.cell(row=1, column=3).value
                                                                                                                    # print(res)

                                                                                                                    def main(i,j):
                                                                                                                            res = wb[wb.get_sheet_names()[0]].cell(row=i, column=j).value
                                                                                                                                return res

                                                                                                                            def translate(content):
                                                                                                                                 
                                                                                                                                    Request_URL = 'http://fanyi.youdao.com/translate?smartresult=dict&smartresult=rule'
                                                                                                                                        Form_Data = {}
                                                                                                                                            Form_Data['i'] = content
                                                                                                                                                Form_Data['from'] = 'AUTO'
                                                                                                                                                    Form_Data['to'] = 'AUTO'
                                                                                                                                                        Form_Data['smartresult'] = 'dict'
                                                                                                                                                            Form_Data['client'] = 'fanyideskweb'
                                                                                                                                                                Form_Data['doctype'] = 'json'
                                                                                                                                                                    Form_Data['version'] = '2.1'
                                                                                                                                                                        Form_Data['keyfrom'] = 'fanyi.web'
                                                                                                                                                                            Form_Data['action'] = 'FY_BY_REALTIME'
                                                                                                                                                                                Form_Data['typoResult'] = 'false'
                                                                                                                                                                                    data = parse.urlencode(Form_Data).encode('utf-8')
                                                                                                                                                                                        response = request.urlopen(Request_URL, data)
                                                                                                                                                                                            html = response.read().decode('utf-8')
                                                                                                                                                                                                translate_results = json.loads(html)
                                                                                                                                                                                                    translate_results = translate_results['translateResult'][0][0]['tgt']
                                                                                                                                                                                                        return  translate_results

                                                                                                                                                                                                    if __name__ == '__main__':
                                                                                                                                                                                                            for i in range(1,226):
                                                                                                                                                                                                                        for j in range(1,15):
                                                                                                                                                                                                                                        print(translate(main(i,j)))
